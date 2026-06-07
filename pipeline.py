"""
Pipeline: select a JSON from input/, extract fields, detect type, transform, save output.
"""
import json
import sys
from pathlib import Path

from extract_fields import process_file
from transform_international import (
    parse_zoning_matrix, build_country_regions, write_country_regions_txt,
    parse_main_costs, build_lanes, write_output as write_international_output,
    is_ma_international, build_regions_ma, write_regions_ma_txt,
    parse_main_costs_ma, write_ma_output,
    CARRIER_MAP,
)
from transform_domestic import (
    check_zones_identical, parse_domestic_at, build_domestic_brackets_at,
    parse_domestic_ro, build_domestic_brackets_ro_express,
    build_domestic_brackets_ro_economy, parse_domestic_hu, filter_weights_hu,
    parse_domestic_de, parse_domestic_pl, build_domestic_brackets_pl,
    parse_domestic_pt, build_postal_code_zones_pt_lines,
    PREMIUM_SERVICE_MAP, PT_DEST_ZONES, PT_ORIGIN,
    write_domestic_output,
)
from transform_accessorials import (
    parse_international_accessorials, parse_domestic_accessorials,
    add_accessorial_sheet, COLUMNS,
)
from openpyxl import load_workbook
from format_xlsx import format_all_output

INPUT_DIR = Path("input")
PROCESSING_DIR = Path("processing")
OUTPUT_DIR = Path("output")

PROCESSING_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)


def select_file() -> Path:
    """List JSON files in input/ and let user pick one."""
    json_files = sorted(INPUT_DIR.glob("*.json"))
    if not json_files:
        print("No JSON files found in input/")
        sys.exit(1)

    print("Available JSON files in input/:")
    for i, f in enumerate(json_files, 1):
        size_mb = f.stat().st_size / (1024 * 1024)
        print(f"  [{i}] {f.name} ({size_mb:.1f} MB)")

    while True:
        choice = input(f"\nSelect file (1-{len(json_files)}): ").strip()
        if choice.isdigit() and 1 <= int(choice) <= len(json_files):
            return json_files[int(choice) - 1]
        print("Invalid choice, try again.")


def extract(input_path: Path) -> Path:
    """Run extract_fields on the selected JSON, save to processing/."""
    print(f"\n--- Extracting fields from: {input_path.name} ---")
    result = process_file(input_path)
    if result is None:
        print("ERROR: No documents found in file.")
        sys.exit(1)

    output_name = input_path.stem + "_extracted.json"
    output_path = PROCESSING_DIR / output_name

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    input_size = input_path.stat().st_size / (1024 * 1024)
    output_size = output_path.stat().st_size / (1024 * 1024)
    print(f"  Input:  {input_size:.2f} MB")
    print(f"  Output: {output_size:.2f} MB -> {output_path}")
    return output_path


def detect_type(fields: dict) -> str:
    """Detect whether the file is international or domestic."""
    if is_ma_international(fields):
        return "international_ma"
    has_zoning = "ZoningMatrix" in fields and bool(fields["ZoningMatrix"].get("value"))
    has_main_costs = "MainCosts" in fields and bool(fields["MainCosts"].get("value"))
    has_domestic = any(
        k.startswith("DomesticMainCosts") and bool(fields[k].get("value"))
        for k in fields
    )

    if has_zoning and has_main_costs:
        return "international"
    if has_domestic:
        return "domestic"
    return "international"


def run_international_ma(fields: dict, base_name: str):
    """Run Morocco (MA) international transform pipeline."""
    print("\n--- Running Morocco International Transform ---")
    out_dir = OUTPUT_DIR

    print("1. Building Country Regions from RegionsMA...")
    regions_data = fields["RegionsMA"]["value"]
    regions = build_regions_ma(regions_data)
    for zone, countries in regions.items():
        print(f"  Zone {zone}: {len(countries)} countries")
    cr_path = out_dir / f"{base_name}_CountryRegions.txt"
    write_regions_ma_txt(regions, cr_path)

    print("2. Parsing MainCostsMA...")
    mc_data = fields["MainCostsMA"]["value"]
    lanes, cost_blocks = parse_main_costs_ma(mc_data)
    for block in cost_blocks:
        print(f"  {block['cost_name']}: {len(block['labels'])} brackets")
    print(f"  Total lanes: {len(lanes)}")

    print("3. Writing output...")
    write_ma_output(lanes, cost_blocks, out_dir)

    default_xlsx = out_dir / "international_rates.xlsx"
    target_xlsx = out_dir / f"{base_name}_rates.xlsx"
    if default_xlsx.exists():
        if target_xlsx.exists():
            target_xlsx.unlink()
        default_xlsx.rename(target_xlsx)
        print(f"  Saved: {target_xlsx}")

    if "AccessorialCosts" in fields and fields["AccessorialCosts"].get("value"):
        print("4. Adding Accessorials...")
        acc_data = fields["AccessorialCosts"]["value"]
        intl_rows = parse_international_accessorials(acc_data)
        wb = load_workbook(target_xlsx)
        if "Accessorials" in wb.sheetnames:
            del wb["Accessorials"]
        add_accessorial_sheet(wb, "Accessorials", intl_rows)
        wb.save(target_xlsx)
        print(f"  Added 'Accessorials' tab ({len(intl_rows)} entries)")

    print("\nMorocco international transform complete!")


def run_international(fields: dict, base_name: str):
    """Run international transform pipeline."""
    print("\n--- Running International Transform ---")
    out_dir = OUTPUT_DIR

    # 1. Parse ZoningMatrix
    print("1. Parsing ZoningMatrix...")
    zm_data = fields["ZoningMatrix"]["value"]
    zone_map, full_zone_map, all_zones = parse_zoning_matrix(zm_data)
    for country in sorted(all_zones.keys()):
        _, code = CARRIER_MAP[country]
        print(f"  {code}: zones {all_zones[country]}")

    # 2. Country Regions
    print("2. Building Country Regions...")
    regions = build_country_regions(full_zone_map)
    cr_path = out_dir / f"{base_name}_CountryRegions.txt"
    write_country_regions_txt(regions, cr_path)

    # 3. Parse MainCosts
    print("3. Parsing MainCosts...")
    mc_data = fields["MainCosts"]["value"]
    rate_blocks = parse_main_costs(mc_data)
    for name, block in rate_blocks.items():
        print(f"  {name}: {len(block['weights'])} weight entries, "
              f"{len(block['rates'])} zones with data")

    # 4. Build lanes
    print("4. Building lanes...")
    lanes = build_lanes(zone_map, all_zones)
    print(f"  Total lanes: {len(lanes)}")

    # 5. Write output
    print("5. Writing output...")
    write_international_output(lanes, rate_blocks, out_dir)

    # Rename output file to include base_name
    default_xlsx = out_dir / "international_rates.xlsx"
    target_xlsx = out_dir / f"{base_name}_rates.xlsx"
    if default_xlsx.exists():
        if target_xlsx.exists():
            target_xlsx.unlink()
        default_xlsx.rename(target_xlsx)
        print(f"  Saved: {target_xlsx}")

    # 6. Accessorials
    if "AccessorialCosts" in fields and fields["AccessorialCosts"].get("value"):
        print("6. Adding Accessorials...")
        acc_data = fields["AccessorialCosts"]["value"]
        intl_rows = parse_international_accessorials(acc_data)
        wb = load_workbook(target_xlsx)
        if "Accessorials" in wb.sheetnames:
            del wb["Accessorials"]
        add_accessorial_sheet(wb, "Accessorials", intl_rows)
        wb.save(target_xlsx)
        print(f"  Added 'Accessorials' tab ({len(intl_rows)} entries)")

    print("\nInternational transform complete!")


def run_domestic(fields: dict, base_name: str):
    """Run domestic transform pipeline."""
    print("\n--- Running Domestic Transform ---")
    out_dir = OUTPUT_DIR
    sheets = []

    # --- DomesticMainCostsAT ---
    if "DomesticMainCostsAT" in fields and fields["DomesticMainCostsAT"].get("value"):
        print("Processing DomesticMainCostsAT...")
        at_data = fields["DomesticMainCostsAT"]["value"]
        zones_same = check_zones_identical(at_data)
        parsed = parse_domestic_at(at_data)
        brackets = build_domestic_brackets_at(parsed["weights"])

        if zones_same:
            header_cols = ["Lane #", "Carrier Name", "Origin Country", "Destination"]
            lane_values = [1, "TNT AT", "AT", "AT"]
        else:
            header_cols = ["Lane #", "Carrier Name", "Origin Country",
                           "Origin Country Region", "Destination",
                           "Destination Country Region"]
            lane_values = [1, "TNT AT", "AT", "", "AT", ""]

        sheets.append({
            "sheet_name": "Domestic Express AT",
            "cost_blocks": [{
                "cost_name": "Transport cost (Domestic Express AT)",
                "brackets": brackets,
                "header_cols": header_cols,
                "lanes": [{"label_values": lane_values, "rates": parsed["rates"]}],
            }],
        })
        print(f"  AT: {len(brackets)} brackets")

    # --- DomesticMainCostsRO ---
    if "DomesticMainCostsRO" in fields and fields["DomesticMainCostsRO"].get("value"):
        print("Processing DomesticMainCostsRO...")
        ro_data = fields["DomesticMainCostsRO"]["value"]
        ro_parsed = parse_domestic_ro(ro_data)
        express_brackets = build_domestic_brackets_ro_express(ro_parsed["express_weights"])
        economy_brackets = build_domestic_brackets_ro_economy(ro_parsed["economy_weights"])

        service_map = ro_parsed["service_map"]
        combined_lanes = []
        lane_num = 1
        for zone_key in ["Zone1", "Zone2", "Zone3", "Zone4"]:
            combined_lanes.append({
                "label_values": [lane_num, service_map[zone_key]],
                "block_rates": [ro_parsed["express_rates"][zone_key], None],
            })
            lane_num += 1
        combined_lanes.append({
            "label_values": [lane_num, "EXP_EC"],
            "block_rates": [None, ro_parsed["economy_rates"]],
        })

        sheets.append({
            "sheet_name": "Domestic Express RO",
            "combined": True,
            "header_cols": ["Lane #", "Service"],
            "cost_blocks": [
                {"cost_name": "Transport cost (Domestic Express RO)", "brackets": express_brackets},
                {"cost_name": "Transport cost (Domestic Economy RO)", "brackets": economy_brackets},
            ],
            "lanes": combined_lanes,
        })
        print(f"  RO: Express {len(express_brackets)} brackets, Economy {len(economy_brackets)} brackets")

    # --- DomesticMainCostsHU ---
    if "DomesticMainCostsHU" in fields and fields["DomesticMainCostsHU"].get("value"):
        print("Processing DomesticMainCostsHU...")
        hu_data = fields["DomesticMainCostsHU"]["value"]
        hu_parsed = parse_domestic_hu(hu_data)
        hu_brackets, hu_z1, hu_z2 = filter_weights_hu(
            hu_parsed["weights"], hu_parsed["rates_z1"], hu_parsed["rates_z2"]
        )

        hu_lanes = [
            {"label_values": [1, "HU", "HU Zone 1", "HU Zone 1"], "block_rates": [hu_z1]},
            {"label_values": [2, "HU", "", "HU Zone 2"], "block_rates": [hu_z2]},
        ]
        sheets.append({
            "sheet_name": "Domestic Express HU",
            "combined": True,
            "header_cols": ["Lane #", "Origin Country", "Origin Postal Code Zone",
                            "Destination Postal Code Zone"],
            "cost_blocks": [{"cost_name": "Transport cost (Domestic Express HU)", "brackets": hu_brackets}],
            "lanes": hu_lanes,
        })

        pcz_lines = [
            "Name of Zone\tCountry\tPostal Code\tExcluded",
            "HU Zone 1\tHU\t109\t",
            "HU Zone 2\tHU\t0-9\t109",
        ]
        pcz_path = out_dir / f"{base_name}_PostalCodeZones_HU.txt"
        pcz_path.write_text("\n".join(pcz_lines), encoding="utf-8")
        print(f"  HU: {len(hu_brackets)} brackets, PostalCodeZones -> {pcz_path.name}")

    # --- DomesticMainCostsDE ---
    if "DomesticMainCostsDE" in fields and fields["DomesticMainCostsDE"].get("value"):
        print("Processing DomesticMainCostsDE...")
        de_data = fields["DomesticMainCostsDE"]["value"]
        de_parsed = parse_domestic_de(de_data)

        de_lanes = []
        lane_num = 1
        for zone_key in ["Zone1", "Zone2", "Zone3", "Zone4", "Zone5"]:
            service = de_parsed["express_service_map"][zone_key]
            de_lanes.append({
                "label_values": [lane_num, "DE", "DE", service],
                "block_rates": [de_parsed["express_rates"][zone_key], None],
            })
            lane_num += 1
        for zone_key in ["Zone1", "Zone2", "Zone3", "Zone4"]:
            service = de_parsed["bpak_service_map"][zone_key]
            de_lanes.append({
                "label_values": [lane_num, "DE", "DE", service],
                "block_rates": [None, de_parsed["bpak_rates"][zone_key]],
            })
            lane_num += 1

        sheets.append({
            "sheet_name": "Domestic Express DE",
            "combined": True,
            "header_cols": ["Lane #", "Origin Country", "Destination Country", "Service"],
            "cost_blocks": [
                {"cost_name": "Transport cost (Domestic Express DE)", "brackets": de_parsed["express_brackets"]},
                {"cost_name": "Transport cost (Domestic Business Pak DE)", "brackets": de_parsed["bpak_brackets"]},
            ],
            "lanes": de_lanes,
        })
        print(f"  DE: Express {len(de_parsed['express_brackets'])} + BPak {len(de_parsed['bpak_brackets'])} brackets")

    # --- DomesticMainCostsPL ---
    if "DomesticMainCostsPL" in fields and fields["DomesticMainCostsPL"].get("value"):
        print("Processing DomesticMainCostsPL...")
        pl_data = fields["DomesticMainCostsPL"]["value"]
        pl_parsed = parse_domestic_pl(pl_data)

        pl_services_order = ["EXP", "EXP1200", "EXP1000", "EXP900"]
        pl_all_brackets = None
        pl_lanes = []
        lane_num = 1
        for svc in pl_services_order:
            if svc not in pl_parsed["services"]:
                continue
            entries = pl_parsed["services"][svc]
            brackets, rates = build_domestic_brackets_pl(entries)
            if pl_all_brackets is None:
                pl_all_brackets = brackets
            while len(rates) < len(pl_all_brackets):
                rates.append("")
            pl_lanes.append({
                "label_values": [lane_num, "PL", "", "PL", "", svc],
                "block_rates": [rates],
            })
            lane_num += 1

        if pl_all_brackets:
            sheets.append({
                "sheet_name": "Domestic Express PL",
                "combined": True,
                "header_cols": ["Lane #", "Origin Country", "Origin Postal Code",
                                "Destination Country", "Destination Postal Code", "Service"],
                "currency": "PLN",
                "cost_blocks": [{"cost_name": "Transport cost (Parcels/Documents)", "brackets": pl_all_brackets}],
                "lanes": pl_lanes,
            })
            print(f"  PL: {len(pl_all_brackets)} brackets, {len(pl_lanes)} services")

    # --- DomesticMainCostsPT ---
    if "DomesticMainCostsPT" in fields and fields["DomesticMainCostsPT"].get("value"):
        print("Processing DomesticMainCostsPT...")
        pt_data = fields["DomesticMainCostsPT"]["value"]
        pt_parsed = parse_domestic_pt(pt_data)

        pt_cost_blocks = [
            {"cost_name": "Transport cost", "brackets": pt_parsed["express_brackets"]},
        ]
        for svc_key, (cost_name, applies_if) in PREMIUM_SERVICE_MAP.items():
            if svc_key in pt_parsed["premium_services"]:
                pt_cost_blocks.append({
                    "cost_name": cost_name,
                    "brackets": pt_parsed["premium_brackets"],
                    "applies_if": applies_if,
                })

        pt_lanes = []
        lane_num = 1
        for zone_key, dest_label in PT_DEST_ZONES:
            block_rates = [pt_parsed["express_rates"][zone_key]]
            for svc_key in PREMIUM_SERVICE_MAP:
                if svc_key in pt_parsed["premium_services"]:
                    rates = list(pt_parsed["premium_services"][svc_key])
                    while len(rates) < 2:
                        rates.append("")
                    block_rates.append(rates)
                else:
                    block_rates.append(None)
            pt_lanes.append({
                "label_values": [lane_num, PT_ORIGIN, dest_label],
                "block_rates": block_rates,
            })
            lane_num += 1

        sheets.append({
            "sheet_name": "Domestic Express PT",
            "combined": True,
            "header_cols": ["Lane #", "Origin Postal Code Zone", "Destination Postal Code Zone"],
            "cost_blocks": pt_cost_blocks,
            "lanes": pt_lanes,
        })

        pcz_path = out_dir / f"{base_name}_PostalCodeZones_PT.txt"
        pcz_path.write_text("\n".join(build_postal_code_zones_pt_lines()), encoding="utf-8")
        print(f"  PT: {len(pt_lanes)} lanes, PostalCodeZones -> {pcz_path.name}")

    # Write domestic xlsx
    if sheets:
        write_domestic_output(sheets, out_dir)
        default_xlsx = out_dir / "domestic_rates.xlsx"
        target_xlsx = out_dir / f"{base_name}_rates.xlsx"
        if default_xlsx.exists():
            if target_xlsx.exists():
                target_xlsx.unlink()
            default_xlsx.rename(target_xlsx)
            print(f"  Saved: {target_xlsx}")

        # Accessorials
        if "AccessorialCosts" in fields and fields["AccessorialCosts"].get("value"):
            print("Adding Accessorials...")
            acc_data = fields["AccessorialCosts"]["value"]
            domestic_by_country = parse_domestic_accessorials(acc_data)
            wb = load_workbook(target_xlsx)
            country_order = ["HU", "RO", "DE", "PL", "PT", "Non-defined"]
            for country in country_order:
                rows = domestic_by_country.get(country, [])
                if not rows:
                    continue
                sheet_name = f"Acc {country}"
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]
                add_accessorial_sheet(wb, sheet_name, rows)
                print(f"  Added 'Acc {country}' ({len(rows)} entries)")
            wb.save(target_xlsx)

    print("\nDomestic transform complete!")


def main():
    input_path = select_file()
    extracted_path = extract(input_path)

    # Load extracted data
    with open(extracted_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    fields = data["documents"][0]["fields"]

    # Detect type and run appropriate transform
    file_type = detect_type(fields)
    base_name = input_path.stem.replace(".pdf", "").replace(".json", "").strip()
    print(f"\nDetected type: {file_type.upper()}")

    if file_type == "international_ma":
        run_international_ma(fields, base_name)
    elif file_type == "international":
        run_international(fields, base_name)
    else:
        run_domestic(fields, base_name)

    # Format output
    print("\n--- Formatting output ---")
    format_all_output(OUTPUT_DIR)

    print("\n=== Pipeline complete! ===")


if __name__ == "__main__":
    main()
