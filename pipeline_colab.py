"""
Pipeline for Google Colab: mount Drive, select JSON, extract, transform, format.
Run this file directly in Colab via "Open in Colab" or upload to Colab.
"""
import json
import sys
import os
from pathlib import Path

# ==============================================================================
# HARDCODED PATHS - Google Drive (Colab)
# ==============================================================================
HARDCODED_INPUT_DIR = "/content/drive/Shareddrives/FA Ops Europe: Rate Maintenance Team /Documents/AI Adoption RMT/RMT_APTIV_VERSIGENT/RMT_FedEx/input"
HARDCODED_OUTPUT_DIR = "/content/drive/Shareddrives/FA Ops Europe: Rate Maintenance Team /Documents/AI Adoption RMT/RMT_APTIV_VERSIGENT/RMT_FedEx/output"
HARDCODED_PROCESSING_DIR = "/content/drive/Shareddrives/FA Ops Europe: Rate Maintenance Team /Documents/AI Adoption RMT/RMT_APTIV_VERSIGENT/RMT_FedEx/processing"
# ==============================================================================


SCRIPTS_DIR_COLAB = "/content/transformation-fedex-aptiv"


def setup_colab():
    """Mount Google Drive and install dependencies."""
    from google.colab import drive
    drive.mount("/content/drive")

    try:
        import openpyxl
    except ImportError:
        os.system("pip install openpyxl -q")

    if SCRIPTS_DIR_COLAB not in sys.path:
        sys.path.insert(0, SCRIPTS_DIR_COLAB)


def main():
    setup_colab()

    # Set paths
    INPUT_DIR = Path(HARDCODED_INPUT_DIR)
    PROCESSING_DIR = Path(HARDCODED_PROCESSING_DIR)
    OUTPUT_DIR = Path(HARDCODED_OUTPUT_DIR)

    PROCESSING_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # Import modules after path setup
    from extract_fields import process_file
    from transform_international import (
        parse_zoning_matrix, build_country_regions, write_country_regions_txt,
        parse_main_costs, build_lanes, write_output as write_international_output,
        is_ma_international, build_regions_ma, write_regions_ma_txt,
        parse_main_costs_ma, write_ma_output,
        CARRIER_MAP,
    )
    from transform_domestic import (
        check_zones_identical, parse_domestic_at, build_domestic_brackets_at,
        parse_domestic_ro, build_domestic_brackets_ro_express,
        build_domestic_brackets_ro_economy, parse_domestic_hu, filter_weights_hu,
        parse_domestic_de, parse_domestic_pl, build_domestic_brackets_pl,
        parse_domestic_pt, build_postal_code_zones_pt_lines,
        PREMIUM_SERVICE_MAP, PT_DEST_ZONES, PT_ORIGIN,
        write_domestic_output,
    )
    from transform_accessorials import (
        parse_international_accessorials, parse_domestic_accessorials,
        add_accessorial_sheet,
    )
    from format_xlsx import format_all_output
    from openpyxl import load_workbook

    # --- Select file ---
    json_files = sorted(INPUT_DIR.glob("*.json"))
    if not json_files:
        print(f"No JSON files found in: {INPUT_DIR}")
        return

    print("\nAvailable JSON files:")
    for i, f in enumerate(json_files, 1):
        size_mb = f.stat().st_size / (1024 * 1024)
        print(f"  [{i}] {f.name} ({size_mb:.1f} MB)")

    choice = input(f"\nSelect file (1-{len(json_files)}): ").strip()
    if not choice.isdigit() or not (1 <= int(choice) <= len(json_files)):
        print("Invalid choice.")
        return
    input_path = json_files[int(choice) - 1]

    # --- Extract fields ---
    print(f"\n--- Extracting fields from: {input_path.name} ---")
    result = process_file(input_path)
    if result is None:
        print("ERROR: No documents found in file.")
        return

    output_name = input_path.stem + "_extracted.json"
    extracted_path = PROCESSING_DIR / output_name
    with open(extracted_path, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    input_size = input_path.stat().st_size / (1024 * 1024)
    output_size = extracted_path.stat().st_size / (1024 * 1024)
    print(f"  Input:  {input_size:.2f} MB")
    print(f"  Output: {output_size:.2f} MB -> {extracted_path.name}")

    # --- Load and detect type ---
    with open(extracted_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    fields = data["documents"][0]["fields"]
    base_name = input_path.stem.replace(".pdf", "").replace(".json", "").strip()

    has_zoning = "ZoningMatrix" in fields and bool(fields["ZoningMatrix"].get("value"))
    has_main_costs = "MainCosts" in fields and bool(fields["MainCosts"].get("value"))
    has_domestic = any(
        k.startswith("DomesticMainCosts") and bool(fields[k].get("value"))
        for k in fields
    )

    if is_ma_international(fields):
        file_type = "international_ma"
    elif has_zoning and has_main_costs:
        file_type = "international"
    elif has_domestic:
        file_type = "domestic"
    else:
        file_type = "international"

    print(f"\nDetected type: {file_type.upper()}")

    # --- MOROCCO INTERNATIONAL ---
    if file_type == "international_ma":
        print("\n--- Running Morocco International Transform ---")

        print("1. Building Country Regions from RegionsMA...")
        regions_data = fields["RegionsMA"]["value"]
        regions = build_regions_ma(regions_data)
        cr_path = OUTPUT_DIR / f"{base_name}_CountryRegions.txt"
        write_regions_ma_txt(regions, cr_path)

        print("2. Parsing MainCostsMA...")
        mc_data = fields["MainCostsMA"]["value"]
        lanes, cost_blocks = parse_main_costs_ma(mc_data)
        print(f"  Total lanes: {len(lanes)}")

        print("3. Writing output...")
        write_ma_output(lanes, cost_blocks, OUTPUT_DIR)

        default_xlsx = OUTPUT_DIR / "international_rates.xlsx"
        target_xlsx = OUTPUT_DIR / f"{base_name}_rates.xlsx"
        if default_xlsx.exists():
            if target_xlsx.exists():
                target_xlsx.unlink()
            default_xlsx.rename(target_xlsx)
            print(f"  Saved: {target_xlsx.name}")

        if "AccessorialCosts" in fields and fields["AccessorialCosts"].get("value"):
            print("4. Adding Accessorials...")
            acc_data = fields["AccessorialCosts"]["value"]
            intl_rows = parse_international_accessorials(acc_data)
            wb = load_workbook(target_xlsx)
            if "Accessorials" in wb.sheetnames:
                del wb["Accessorials"]
            add_accessorial_sheet(wb, "Accessorials", intl_rows)
            wb.save(target_xlsx)
            print(f"  Added 'Accessorials' tab ({len(intl_rows)} entries)")

    # --- INTERNATIONAL ---
    elif file_type == "international":
        print("\n--- Running International Transform ---")

        print("1. Parsing ZoningMatrix...")
        zm_data = fields["ZoningMatrix"]["value"]
        zone_map, full_zone_map, all_zones = parse_zoning_matrix(zm_data)
        for country in sorted(all_zones.keys()):
            _, code = CARRIER_MAP[country]
            print(f"  {code}: zones {all_zones[country]}")

        print("2. Building Country Regions...")
        regions = build_country_regions(full_zone_map)
        cr_path = OUTPUT_DIR / f"{base_name}_CountryRegions.txt"
        write_country_regions_txt(regions, cr_path)

        print("3. Parsing MainCosts...")
        mc_data = fields["MainCosts"]["value"]
        rate_blocks = parse_main_costs(mc_data)
        for name, block in rate_blocks.items():
            print(f"  {name}: {len(block['weights'])} weights, {len(block['rates'])} zones")

        print("4. Building lanes...")
        lanes = build_lanes(zone_map, all_zones)
        print(f"  Total lanes: {len(lanes)}")

        print("5. Writing output...")
        write_international_output(lanes, rate_blocks, OUTPUT_DIR)

        default_xlsx = OUTPUT_DIR / "international_rates.xlsx"
        target_xlsx = OUTPUT_DIR / f"{base_name}_rates.xlsx"
        if default_xlsx.exists():
            if target_xlsx.exists():
                target_xlsx.unlink()
            default_xlsx.rename(target_xlsx)
            print(f"  Saved: {target_xlsx.name}")

        if "AccessorialCosts" in fields and fields["AccessorialCosts"].get("value"):
            print("6. Adding Accessorials...")
            acc_data = fields["AccessorialCosts"]["value"]
            intl_rows = parse_international_accessorials(acc_data)
            wb = load_workbook(target_xlsx)
            if "Accessorials" in wb.sheetnames:
                del wb["Accessorials"]
            add_accessorial_sheet(wb, "Accessorials", intl_rows)
            wb.save(target_xlsx)
            print(f"  Added 'Accessorials' tab ({len(intl_rows)} entries)")

    # --- DOMESTIC ---
    else:
        print("\n--- Running Domestic Transform ---")
        sheets = []

        if "DomesticMainCostsAT" in fields and fields["DomesticMainCostsAT"].get("value"):
            print("Processing DomesticMainCostsAT...")
            at_data = fields["DomesticMainCostsAT"]["value"]
            zones_same = check_zones_identical(at_data)
            parsed = parse_domestic_at(at_data)
            brackets = build_domestic_brackets_at(parsed["weights"])
            if zones_same:
                header_cols = ["Lane #", "Carrier Name", "Origin Country", "Destination"]
                lane_values = [1, "TNT AT", "AT", "AT"]
            else:
                header_cols = ["Lane #", "Carrier Name", "Origin Country",
                               "Origin Country Region", "Destination", "Destination Country Region"]
                lane_values = [1, "TNT AT", "AT", "", "AT", ""]
            sheets.append({
                "sheet_name": "Domestic Express AT",
                "cost_blocks": [{"cost_name": "Transport cost (Domestic Express AT)",
                                 "brackets": brackets, "header_cols": header_cols,
                                 "lanes": [{"label_values": lane_values, "rates": parsed["rates"]}]}],
            })
            print(f"  AT: {len(brackets)} brackets")

        if "DomesticMainCostsRO" in fields and fields["DomesticMainCostsRO"].get("value"):
            print("Processing DomesticMainCostsRO...")
            ro_data = fields["DomesticMainCostsRO"]["value"]
            ro_parsed = parse_domestic_ro(ro_data)
            express_brackets = build_domestic_brackets_ro_express(ro_parsed["express_weights"])
            economy_brackets = build_domestic_brackets_ro_economy(ro_parsed["economy_weights"])
            service_map = ro_parsed["service_map"]
            combined_lanes = []
            lane_num = 1
            for zone_key in ["Zone1", "Zone2", "Zone3", "Zone4"]:
                combined_lanes.append({"label_values": [lane_num, service_map[zone_key]],
                                       "block_rates": [ro_parsed["express_rates"][zone_key], None]})
                lane_num += 1
            combined_lanes.append({"label_values": [lane_num, "EXP_EC"],
                                   "block_rates": [None, ro_parsed["economy_rates"]]})
            sheets.append({
                "sheet_name": "Domestic Express RO", "combined": True,
                "header_cols": ["Lane #", "Service"],
                "cost_blocks": [
                    {"cost_name": "Transport cost (Domestic Express RO)", "brackets": express_brackets},
                    {"cost_name": "Transport cost (Domestic Economy RO)", "brackets": economy_brackets},
                ], "lanes": combined_lanes,
            })
            print(f"  RO: Express {len(express_brackets)}, Economy {len(economy_brackets)} brackets")

        if "DomesticMainCostsHU" in fields and fields["DomesticMainCostsHU"].get("value"):
            print("Processing DomesticMainCostsHU...")
            hu_data = fields["DomesticMainCostsHU"]["value"]
            hu_parsed = parse_domestic_hu(hu_data)
            hu_brackets, hu_z1, hu_z2 = filter_weights_hu(
                hu_parsed["weights"], hu_parsed["rates_z1"], hu_parsed["rates_z2"])
            hu_lanes = [
                {"label_values": [1, "HU", "HU Zone 1", "HU Zone 1"], "block_rates": [hu_z1]},
                {"label_values": [2, "HU", "", "HU Zone 2"], "block_rates": [hu_z2]},
            ]
            sheets.append({
                "sheet_name": "Domestic Express HU", "combined": True,
                "header_cols": ["Lane #", "Origin Country", "Origin Postal Code Zone",
                                "Destination Postal Code Zone"],
                "cost_blocks": [{"cost_name": "Transport cost (Domestic Express HU)", "brackets": hu_brackets}],
                "lanes": hu_lanes,
            })
            pcz_lines = ["Name of Zone\tCountry\tPostal Code\tExcluded",
                         "HU Zone 1\tHU\t109\t", "HU Zone 2\tHU\t0-9\t109"]
            pcz_path = OUTPUT_DIR / f"{base_name}_PostalCodeZones_HU.txt"
            pcz_path.write_text("\n".join(pcz_lines), encoding="utf-8")
            print(f"  HU: {len(hu_brackets)} brackets")

        if "DomesticMainCostsDE" in fields and fields["DomesticMainCostsDE"].get("value"):
            print("Processing DomesticMainCostsDE...")
            de_data = fields["DomesticMainCostsDE"]["value"]
            de_parsed = parse_domestic_de(de_data)
            de_lanes = []
            lane_num = 1
            for zone_key in ["Zone1", "Zone2", "Zone3", "Zone4", "Zone5"]:
                service = de_parsed["express_service_map"][zone_key]
                de_lanes.append({"label_values": [lane_num, "DE", "DE", service],
                                 "block_rates": [de_parsed["express_rates"][zone_key], None]})
                lane_num += 1
            for zone_key in ["Zone1", "Zone2", "Zone3", "Zone4"]:
                service = de_parsed["bpak_service_map"][zone_key]
                de_lanes.append({"label_values": [lane_num, "DE", "DE", service],
                                 "block_rates": [None, de_parsed["bpak_rates"][zone_key]]})
                lane_num += 1
            sheets.append({
                "sheet_name": "Domestic Express DE", "combined": True,
                "header_cols": ["Lane #", "Origin Country", "Destination Country", "Service"],
                "cost_blocks": [
                    {"cost_name": "Transport cost (Domestic Express DE)", "brackets": de_parsed["express_brackets"]},
                    {"cost_name": "Transport cost (Domestic Business Pak DE)", "brackets": de_parsed["bpak_brackets"]},
                ], "lanes": de_lanes,
            })
            print(f"  DE: Express {len(de_parsed['express_brackets'])} + BPak {len(de_parsed['bpak_brackets'])}")

        if "DomesticMainCostsPL" in fields and fields["DomesticMainCostsPL"].get("value"):
            print("Processing DomesticMainCostsPL...")
            pl_data = fields["DomesticMainCostsPL"]["value"]
            pl_parsed = parse_domestic_pl(pl_data)
            pl_services_order = ["EXP", "EXP1200", "EXP1000", "EXP900"]
            pl_all_brackets = None
            pl_lanes = []
            lane_num = 1
            for svc in pl_services_order:
                if svc not in pl_parsed["services"]:
                    continue
                entries = pl_parsed["services"][svc]
                brackets, rates = build_domestic_brackets_pl(entries)
                if pl_all_brackets is None:
                    pl_all_brackets = brackets
                while len(rates) < len(pl_all_brackets):
                    rates.append("")
                pl_lanes.append({"label_values": [lane_num, "PL", "", "PL", "", svc],
                                 "block_rates": [rates]})
                lane_num += 1
            if pl_all_brackets:
                sheets.append({
                    "sheet_name": "Domestic Express PL", "combined": True,
                    "header_cols": ["Lane #", "Origin Country", "Origin Postal Code",
                                    "Destination Country", "Destination Postal Code", "Service"],
                    "currency": "PLN",
                    "cost_blocks": [{"cost_name": "Transport cost (Parcels/Documents)", "brackets": pl_all_brackets}],
                    "lanes": pl_lanes,
                })
                print(f"  PL: {len(pl_all_brackets)} brackets, {len(pl_lanes)} services")

        if "DomesticMainCostsPT" in fields and fields["DomesticMainCostsPT"].get("value"):
            print("Processing DomesticMainCostsPT...")
            pt_data = fields["DomesticMainCostsPT"]["value"]
            pt_parsed = parse_domestic_pt(pt_data)
            pt_cost_blocks = [
                {"cost_name": "Transport cost", "brackets": pt_parsed["express_brackets"]},
            ]
            for svc_key, (cost_name, applies_if) in PREMIUM_SERVICE_MAP.items():
                if svc_key in pt_parsed["premium_services"]:
                    pt_cost_blocks.append({
                        "cost_name": cost_name,
                        "brackets": pt_parsed["premium_brackets"],
                        "applies_if": applies_if,
                    })
            pt_lanes = []
            lane_num = 1
            for zone_key, dest_label in PT_DEST_ZONES:
                block_rates = [pt_parsed["express_rates"][zone_key]]
                for svc_key in PREMIUM_SERVICE_MAP:
                    if svc_key in pt_parsed["premium_services"]:
                        rates = list(pt_parsed["premium_services"][svc_key])
                        while len(rates) < 2:
                            rates.append("")
                        block_rates.append(rates)
                    else:
                        block_rates.append(None)
                pt_lanes.append({
                    "label_values": [lane_num, PT_ORIGIN, dest_label],
                    "block_rates": block_rates,
                })
                lane_num += 1
            sheets.append({
                "sheet_name": "Domestic Express PT",
                "combined": True,
                "header_cols": ["Lane #", "Origin Postal Code Zone", "Destination Postal Code Zone"],
                "cost_blocks": pt_cost_blocks,
                "lanes": pt_lanes,
            })
            pcz_path = OUTPUT_DIR / f"{base_name}_PostalCodeZones_PT.txt"
            pcz_path.write_text("\n".join(build_postal_code_zones_pt_lines()), encoding="utf-8")
            print(f"  PT: {len(pt_lanes)} lanes")

        if sheets:
            write_domestic_output(sheets, OUTPUT_DIR)
            default_xlsx = OUTPUT_DIR / "domestic_rates.xlsx"
            target_xlsx = OUTPUT_DIR / f"{base_name}_rates.xlsx"
            if default_xlsx.exists():
                if target_xlsx.exists():
                    target_xlsx.unlink()
                default_xlsx.rename(target_xlsx)
                print(f"  Saved: {target_xlsx.name}")

            if "AccessorialCosts" in fields and fields["AccessorialCosts"].get("value"):
                print("Adding Accessorials...")
                acc_data = fields["AccessorialCosts"]["value"]
                domestic_by_country = parse_domestic_accessorials(acc_data)
                wb = load_workbook(target_xlsx)
                for country in ["HU", "RO", "DE", "PL", "PT", "Non-defined"]:
                    rows = domestic_by_country.get(country, [])
                    if not rows:
                        continue
                    sn = f"Acc {country}"
                    if sn in wb.sheetnames:
                        del wb[sn]
                    add_accessorial_sheet(wb, sn, rows)
                    print(f"  Added 'Acc {country}' ({len(rows)} entries)")
                wb.save(target_xlsx)

    # --- Format output ---
    print("\n--- Formatting output ---")
    format_all_output(OUTPUT_DIR)

    print("\n=== Pipeline complete! ===")
    print(f"Output saved to: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
