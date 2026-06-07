import json
import re
from pathlib import Path
from openpyxl import Workbook

INTL_FILE = Path("processing/fedex international 4.pdf_extracted.json")
DOMESTIC_FILE = Path("processing/fedex domestic 6(with pln).pdf_extracted.json")
OUTPUT_DIR = Path("output")
OUTPUT_DIR.mkdir(exist_ok=True)

COLUMNS = [
    "Rate Card cost Name",
    "Rate Agreement cost name",
    "Cost",
    "Currency",
    "MIN",
    "MAX",
    "Apply if",
    "Rate by",
]


def parse_price_simple(price_str: str) -> dict:
    """Parse simple price strings into cost, currency, min, max, rate_by."""
    result = {"cost": "", "currency": "", "min": "", "max": "", "rate_by": ""}

    if not price_str:
        return result

    price = price_str.strip()

    # Handle "on request", "EXEMPT", descriptive prices
    if price.lower() in ("on request", "exempt"):
        result["cost"] = price
        return result

    # Handle percentage format: "0.25% MIN 8 PLN"
    pct_min_match = re.match(r"([\d.,]+)%\s*MIN\s*([\d.,]+)\s*(\w+)", price, re.IGNORECASE)
    if pct_min_match:
        result["cost"] = f"{pct_min_match.group(1)}%"
        result["min"] = pct_min_match.group(2)
        result["currency"] = pct_min_match.group(3)
        result["rate_by"] = "Value of Goods"
        return result

    # Handle "1% minimum 10 EUR"
    pct_min_match2 = re.match(r"([\d.,]+)%\s*minimum\s*([\d.,]+)\s*(\w+)", price, re.IGNORECASE)
    if pct_min_match2:
        result["cost"] = f"{pct_min_match2.group(1)}%"
        result["min"] = pct_min_match2.group(2)
        result["currency"] = pct_min_match2.group(3)
        result["rate_by"] = "Value of Goods"
        return result

    # Handle pure percentage: "1.40%", "0.00%"
    if re.match(r"^[\d.,]+%$", price):
        result["cost"] = price
        return result

    # Handle "Index -60%" (fuel surcharge)
    if "index" in price.lower() or "discount" in price.lower():
        result["cost"] = price
        return result

    # Handle "Domestic Express price + 100%"
    if "price" in price.lower() and "%" in price:
        result["cost"] = price
        return result

    # Handle "50% of shipment's price"
    if "%" in price and "of" in price.lower():
        result["cost"] = price
        return result

    # Handle "0.35 EUR per kg with minimum of 18.00 EUR per con"
    per_kg_match = re.match(
        r"([\d.,]+)\s*(EUR|PLN|€)\s*per\s*kg.*?minimum.*?([\d.,]+)",
        price, re.IGNORECASE
    )
    if per_kg_match:
        result["cost"] = per_kg_match.group(1)
        result["currency"] = per_kg_match.group(2).replace("€", "EUR")
        result["min"] = per_kg_match.group(3)
        result["rate_by"] = "Weight/chargeable kg"
        return result

    # Handle "€0.35 per kg with minimum of €18.00 per con"
    per_kg_match2 = re.match(
        r"€\s*([\d.,]+)\s*per\s*kg.*?minimum.*?€\s*([\d.,]+)",
        price, re.IGNORECASE
    )
    if per_kg_match2:
        result["cost"] = per_kg_match2.group(1)
        result["currency"] = "EUR"
        result["min"] = per_kg_match2.group(2)
        result["rate_by"] = "Weight/chargeable kg"
        return result

    # Handle "EUR\n35.00" or "EUR\n10.00\n60.00"
    eur_nl_match = re.match(r"(EUR|PLN)\n([\d.,]+)", price)
    if eur_nl_match:
        result["currency"] = eur_nl_match.group(1)
        result["cost"] = eur_nl_match.group(2)
        return result

    # Handle "€ 14.56" or "€14.56"
    euro_match = re.match(r"€\s*([\d.,]+)", price)
    if euro_match:
        result["cost"] = euro_match.group(1)
        result["currency"] = "EUR"
        return result

    # Handle "9.90 €" or "5.00 €"
    euro_suffix_match = re.match(r"([\d.,]+)\s*€", price)
    if euro_suffix_match:
        result["cost"] = euro_suffix_match.group(1)
        result["currency"] = "EUR"
        return result

    # Handle "35.00" (number only)
    num_only = re.match(r"^([\d.,]+)$", price)
    if num_only:
        result["cost"] = num_only.group(1)
        return result

    # Handle "12 PLN" or "55 PLN"
    num_curr = re.match(r"([\d.,]+)\s*(EUR|PLN)", price)
    if num_curr:
        result["cost"] = num_curr.group(1)
        result["currency"] = num_curr.group(2)
        return result

    # Handle "0 PLN (zero)" or "0 EUR (zero)"
    zero_match = re.match(r"([\d.,]+)\s*(EUR|PLN)\s*\(zero\)", price, re.IGNORECASE)
    if zero_match:
        result["cost"] = zero_match.group(1)
        result["currency"] = zero_match.group(2)
        return result

    # Handle "0,00 €"
    comma_euro = re.match(r"([\d,]+)\s*€", price)
    if comma_euro:
        result["cost"] = comma_euro.group(1).replace(",", ".")
        result["currency"] = "EUR"
        return result

    # Fallback
    result["cost"] = price
    return result


def parse_rate_by(rate_by_str: str) -> str:
    """Normalize rate_by string."""
    if not rate_by_str:
        return ""
    rb = rate_by_str.strip().lower()
    if "fee per consignment" in rb or "surcharge per consignment" in rb:
        return "Consignment"
    if "fee per shipment" in rb:
        return "Shipment"
    if "fee per invoice" in rb or "per invoice" in rb:
        return "Invoice"
    if "percentage of declared value" in rb or "% value of goods" in rb or "valor da mercadoria" in rb:
        return "Value of Goods"
    if "additional item" in rb:
        return "Additional item"
    if "number of days" in rb or "número de dias" in rb:
        return "Amount of days"
    if "shipment" in rb or "envio" in rb:
        return "Shipment"
    if "fatura" in rb:
        return "Invoice"
    return rate_by_str.strip()


def determine_apply_if(cost_name: str) -> str:
    """Determine 'Apply if' based on cost name."""
    if "09:00" in cost_name or "9:00" in cost_name:
        return "Service equals EXP900"
    if "10:00" in cost_name:
        return "Service equals EXP1000"
    if "12:00 Express" in cost_name and "Economy" not in cost_name:
        return "Service equals EXP1200"
    if "12:00 Economy" in cost_name:
        return "Service equals Economy EXP1200"
    return ""


def parse_international_accessorials(data: list) -> list:
    """Parse international AccessorialCosts into rows."""
    rows = []
    for entry in data:
        cost_name = entry.get("CostName", "").strip()
        price = entry.get("Price", "").strip()
        rate_by_raw = entry.get("RateBy", "").strip()

        if not cost_name:
            continue

        # Handle Insurance/Enhanced Liability with percentage + minimum
        if rate_by_raw and "Percentage of declared value" in rate_by_raw and "\n" in price:
            parts = price.split("\n")
            rb_parts = rate_by_raw.split("\n")
            pct_val = parts[0] if parts else ""
            min_val = parts[1] if len(parts) > 1 else ""
            rows.append({
                "name": cost_name,
                "cost": f"{pct_val}%",
                "currency": "EUR",
                "min": min_val,
                "max": "",
                "apply_if": determine_apply_if(cost_name),
                "rate_by": "Value of Goods",
            })
            continue

        # Handle "EUR\n10.00\n60.00" with fuel discount
        if "Discount on the TNT fuel index" in rate_by_raw:
            parts = price.split("\n")
            currency = parts[0] if parts else "EUR"
            cost_val = parts[1] if len(parts) > 1 else ""
            rows.append({
                "name": cost_name,
                "cost": cost_val,
                "currency": currency,
                "min": "",
                "max": "",
                "apply_if": determine_apply_if(cost_name),
                "rate_by": "Consignment",
            })
            continue

        # Standard parsing
        parsed = parse_price_simple(price)
        rate_by = parsed["rate_by"] or parse_rate_by(rate_by_raw)

        # Extract currency from RateBy if not found
        if not parsed["currency"] and rate_by_raw:
            if "EUR" in rate_by_raw:
                parsed["currency"] = "EUR"

        rows.append({
            "name": cost_name,
            "cost": parsed["cost"],
            "currency": parsed["currency"],
            "min": parsed["min"],
            "max": parsed["max"],
            "apply_if": determine_apply_if(cost_name),
            "rate_by": rate_by,
        })

    return rows


def parse_pt_ilhas_contin(price: str, rate_by_raw: str) -> list:
    """Parse PT Ilhas/Continent split prices into rows."""
    lines = [ln.strip() for ln in price.split("\n") if ln.strip()]
    # Drop header labels (Ilhas, Contin., etc.)
    value_lines = [ln for ln in lines if ln.lower() not in ("ilhas", "contin.", "continente")]
    costs = []
    for ln in value_lines:
        if ln.upper() == "N/A":
            costs.append("N/A")
            continue
        m = re.search(r"([\d.,]+)\s*€|€\s*([\d.,]+)", ln)
        if m:
            costs.append((m.group(1) or m.group(2)).replace(",", "."))
    if len(costs) < 2:
        return []
    labels = ["Ilhas", "Continente"]
    rate_by = parse_rate_by(rate_by_raw) or "Shipment"
    rows = []
    for i, cost in enumerate(costs[:2]):
        rows.append({
            "name": "",
            "cost": cost,
            "currency": "EUR" if cost != "N/A" else "",
            "min": "",
            "max": "",
            "apply_if": labels[i] if i < len(labels) else "",
            "rate_by": rate_by,
        })
    return rows


def parse_domestic_accessorials(data: list) -> dict:
    """Parse domestic AccessorialCosts, split by country."""
    countries = {"HU": [], "RO": [], "DE": [], "PL": [], "PT": [], "Non-defined": []}
    current_country = None
    pending_pct = None  # Holds a percentage value from a standalone Price entry

    for entry in data:
        cost_name = entry.get("CostName", "").strip()
        price = entry.get("Price", "").strip()
        rate_by_raw = entry.get("RateBy", "").strip()

        if not cost_name and not price:
            continue

        # Country detection markers
        if "DOMESTIC RATES" in cost_name:
            continue

        if "HUNGARY" in cost_name and "ADDITIONAL" in cost_name:
            current_country = "HU"
            continue

        # RO marker embedded in a CostName
        if "RO\nADDITIONAL SERVICES" in cost_name:
            # The part before "RO\n..." is the last HU entry
            parts = cost_name.split("\nRO\nADDITIONAL SERVICES")
            if parts[0].strip() and current_country == "HU":
                parsed = parse_price_simple(price)
                rate_by = parsed["rate_by"] or parse_rate_by(rate_by_raw)
                countries["HU"].append({
                    "name": parts[0].strip(),
                    "cost": parsed["cost"],
                    "currency": parsed["currency"] or "EUR",
                    "min": parsed["min"],
                    "max": parsed["max"],
                    "apply_if": "",
                    "rate_by": rate_by,
                })
            current_country = "RO"
            continue

        # DE detection: after RO, when "Additional terms and conditions" appears
        if current_country == "RO" and "Additional terms and conditions" in cost_name:
            # This entry still belongs to RO (clean up name)
            clean_name = cost_name.replace("\nAdditional terms and conditions", "").strip()
            if price:
                parsed = parse_price_simple(price)
                rate_by = parsed["rate_by"] or parse_rate_by(rate_by_raw)
                countries["RO"].append({
                    "name": clean_name,
                    "cost": parsed["cost"],
                    "currency": parsed["currency"] or "EUR",
                    "min": parsed["min"],
                    "max": parsed["max"],
                    "apply_if": "",
                    "rate_by": rate_by,
                })
            current_country = "DE"
            continue

        # PL detection
        if cost_name == "OPTION/SURCHARGE":
            current_country = "PL"
            continue

        # PT detection
        if "PARA ENTREGAS EM PORTUGAL" in cost_name or "PORTUGAL" in cost_name.upper() and "WRONG ADDRESS" in cost_name.upper():
            clean = cost_name.split("\n(PARA ENTREGAS EM PORTUGAL")[0].strip()
            if clean and price and current_country == "PL":
                parsed = parse_price_simple(price.split("\n")[0])
                countries["PL"].append({
                    "name": clean,
                    "cost": parsed["cost"],
                    "currency": parsed["currency"] or "PLN",
                    "min": parsed["min"],
                    "max": parsed["max"],
                    "apply_if": "",
                    "rate_by": parse_rate_by(rate_by_raw),
                })
            current_country = "PT"
            if price and ("zero" in price.lower() or "exempt" in price.lower()):
                continue
            continue

        if current_country is None:
            continue

        # Skip empty/header entries
        if not price and not rate_by_raw:
            continue

        # Handle PT Ilhas/Continent split prices
        if current_country == "PT" and "ilhas" in price.lower() and "contin" in price.lower():
            ilhas_rows = parse_pt_ilhas_contin(price, rate_by_raw)
            if ilhas_rows:
                clean_name = cost_name.split("\n")[0].strip()
                for row in ilhas_rows:
                    row["name"] = clean_name
                    countries["PT"].append(row)
                continue

        # Handle PT per-hour wait (two tier prices)
        if current_country == "PT" and "por hora" in price.lower():
            euros = re.findall(r"([\d.,]+)\s*€|€\s*([\d.,]+)", price)
            costs = [((m[0] or m[1]).replace(",", ".")) for m in euros if (m[0] or m[1])]
            clean_name = cost_name.split("\n")[0].strip()
            rate_by = "Hour" if "espera" in rate_by_raw.lower() or "hora" in price.lower() else parse_rate_by(rate_by_raw)
            if len(costs) >= 2:
                countries["PT"].append({
                    "name": clean_name, "cost": costs[0], "currency": "EUR",
                    "min": "", "max": "", "apply_if": "Ilhas", "rate_by": rate_by,
                })
                countries["PT"].append({
                    "name": clean_name, "cost": costs[1], "currency": "EUR",
                    "min": "", "max": "", "apply_if": "Continente", "rate_by": rate_by,
                })
            elif costs:
                countries["PT"].append({
                    "name": clean_name, "cost": costs[0], "currency": "EUR",
                    "min": "", "max": "", "apply_if": "", "rate_by": rate_by,
                })
            continue

        # Handle PT Book slot (Portuguese)
        if current_country == "PT" and "book slot" in cost_name.lower() and "/dia" in price:
            weight_prices = re.findall(r"€\s*([\d.,]+)/dia|([\d.,]+)\s*€/dia", price)
            weight_prices = [(m[0] or m[1]).replace(",", ".") for m in weight_prices]
            weight_tiers = ["Weight <30 kg", "Weight <125 kg", "Weight >125 kg"]
            clean_name = cost_name.split("\n")[0].strip()
            for i, wp in enumerate(weight_prices):
                if i < len(weight_tiers):
                    countries["PT"].append({
                        "name": clean_name,
                        "cost": wp.replace(",", "."),
                        "currency": "EUR",
                        "min": "",
                        "max": "",
                        "apply_if": weight_tiers[i],
                        "rate_by": "Amount of days",
                    })
            continue

        # Handle PT Insurance percentage
        if current_country == "PT" and price.strip() == "1%" and "mercadoria" in rate_by_raw.lower():
            countries["PT"].append({
                "name": cost_name.split("\n")[0].strip(),
                "cost": "1%",
                "currency": "EUR",
                "min": "",
                "max": "",
                "apply_if": "Value of Goods",
                "rate_by": "Value of Goods",
            })
            continue

        # Handle "Book slot" multi-weight entries for RO
        if "Book slot" in cost_name and ("EUR/day" in price or "/dia" in price):
            weight_prices = re.findall(r"([\d.,]+)\s*(?:EUR|€)/(?:day|dia)", price)
            if "up to 30 kg" in rate_by_raw.lower() or "até 30 kg" in rate_by_raw.lower():
                weight_tiers = ["Weight <30 kg", "Weight <125 kg", "Weight >125 kg"]
                for i, wp in enumerate(weight_prices):
                    if i < len(weight_tiers):
                        countries[current_country].append({
                            "name": cost_name.split("\n")[0].strip(),
                            "cost": wp,
                            "currency": "EUR",
                            "min": "",
                            "max": "",
                            "apply_if": weight_tiers[i],
                            "rate_by": "Amount of days",
                        })
                continue

        # Handle "1 % of declared value" entries (price without CostName) - save for next entry
        if not cost_name and price and "%" in price and "declared" in price.lower():
            pct_match = re.match(r"([\d.,]+)\s*%", price)
            if pct_match:
                pending_pct = pct_match.group(1)
            continue

        # Handle continuation from pending percentage (e.g., "of shipment, with a min. of €20")
        if pending_pct and ("min" in price.lower() or "value" in rate_by_raw.lower()):
            min_match = re.search(r"€\s*([\d.,]+)", price)
            min_val = min_match.group(1) if min_match else ""
            countries[current_country].append({
                "name": re.sub(r"\d+$", "", cost_name.split("\n")[0]).strip(),
                "cost": f"{pending_pct}%",
                "currency": "EUR",
                "min": min_val,
                "max": "",
                "apply_if": "",
                "rate_by": "Value of Goods",
            })
            pending_pct = None
            continue

        pending_pct = None

        # Parse standard entry
        parsed = parse_price_simple(price)
        rate_by = parsed["rate_by"] or parse_rate_by(rate_by_raw)

        # Determine default currency by country
        if not parsed["currency"]:
            if current_country == "PL":
                parsed["currency"] = "PLN"
            else:
                parsed["currency"] = "EUR"

        # Clean cost name (remove footnote numbers, trailing chars)
        clean_name = re.sub(r"\d+$", "", cost_name.split("\n")[0]).strip().rstrip(",")

        countries[current_country].append({
            "name": clean_name,
            "cost": parsed["cost"],
            "currency": parsed["currency"],
            "min": parsed["min"],
            "max": parsed["max"],
            "apply_if": "",
            "rate_by": rate_by,
        })

    return countries


def add_accessorial_sheet(wb, sheet_name: str, rows: list):
    """Add an accessorial costs sheet to an existing workbook."""
    ws = wb.create_sheet(title=sheet_name)
    ws.append(COLUMNS)
    for row in rows:
        data = [
            row["name"], "", row["cost"], row["currency"],
            row["min"], row["max"], row["apply_if"], row["rate_by"]
        ]
        ws.append(data)
    return ws


def write_accessorials_output(intl_rows: list, domestic_by_country: dict, output_dir: Path):
    """Add accessorial tabs to existing rates files."""
    from openpyxl import load_workbook as load_wb

    # --- International: add tab to international_rates.xlsx ---
    intl_xlsx = output_dir / "international_rates.xlsx"
    if intl_xlsx.exists():
        wb = load_wb(intl_xlsx)
        # Remove existing accessorial sheet if re-running
        if "Accessorials" in wb.sheetnames:
            del wb["Accessorials"]
        add_accessorial_sheet(wb, "Accessorials", intl_rows)
        wb.save(intl_xlsx)
        print(f"  Added 'Accessorials' tab to: {intl_xlsx}")
    else:
        print(f"  WARNING: {intl_xlsx} not found, creating standalone file")
        wb = Workbook()
        wb.remove(wb.active)
        add_accessorial_sheet(wb, "Accessorials", intl_rows)
        wb.save(intl_xlsx)

    # --- Domestic: add tabs to domestic_rates.xlsx ---
    dom_xlsx = output_dir / "domestic_rates.xlsx"
    if dom_xlsx.exists():
        wb = load_wb(dom_xlsx)
    else:
        print(f"  WARNING: {dom_xlsx} not found, creating standalone file")
        wb = Workbook()
        wb.remove(wb.active)

    country_order = ["HU", "RO", "DE", "PL", "PT", "Non-defined"]
    for country in country_order:
        rows = domestic_by_country.get(country, [])
        if not rows:
            continue
        sheet_name = f"Acc {country}"
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        add_accessorial_sheet(wb, sheet_name, rows)

    wb.save(dom_xlsx)
    print(f"  Added accessorial tabs to: {dom_xlsx}")



def main():
    # Parse international
    print("Processing International Accessorial Costs...")
    with open(INTL_FILE, "r", encoding="utf-8") as f:
        intl_data = json.load(f)
    intl_acc = intl_data["documents"][0]["fields"]["AccessorialCosts"]["value"]
    intl_rows = parse_international_accessorials(intl_acc)
    print(f"  Rows: {len(intl_rows)}")

    # Parse domestic
    print("\nProcessing Domestic Accessorial Costs...")
    with open(DOMESTIC_FILE, "r", encoding="utf-8") as f:
        dom_data = json.load(f)
    dom_acc = dom_data["documents"][0]["fields"]["AccessorialCosts"]["value"]
    domestic_by_country = parse_domestic_accessorials(dom_acc)
    for country, rows in domestic_by_country.items():
        print(f"  {country}: {len(rows)} entries")

    # Write output
    print("\nWriting output...")
    write_accessorials_output(intl_rows, domestic_by_country, OUTPUT_DIR)
    print("\nDone!")


if __name__ == "__main__":
    main()
