"""
Apply visual formatting to all output XLSX files.
"""
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter


# Style constants
HEADER_FONT = Font(bold=True, size=10, name="Calibri")
DATA_FONT = Font(size=10, name="Calibri")
COST_NAME_FONT = Font(bold=True, size=10, name="Calibri", color="1F4E79")

HEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="EBF1F8", end_color="EBF1F8", fill_type="solid")
ACC_HEADER_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

BOTTOM_BORDER = Border(
    bottom=Side(style="thin", color="808080"),
)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)


def auto_width(ws, min_width=8, max_width=18):
    """Set column widths based on content, clamped between min and max."""
    for col_idx in range(1, ws.max_column + 1):
        max_len = min_width
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=False):
            cell = row[0]
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_len:
                    max_len = cell_len
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)


def format_rates_sheet(ws):
    """Format a rates sheet (international or domestic)."""
    if ws.max_row < 2:
        return

    # Find the main header row (contains "Lane #")
    header_row = None
    cost_name_rows = []
    subheader_rows = []

    for row_idx in range(1, ws.max_row + 1):
        cell_val = str(ws.cell(row=row_idx, column=1).value or "")
        if cell_val == "Lane #":
            header_row = row_idx
            # Rows above header are cost name / info rows
            for r in range(1, row_idx):
                val = ws.cell(row=r, column=1).value
                if val is None:
                    # Check other columns for cost name
                    for c in range(2, ws.max_column + 1):
                        v = ws.cell(row=r, column=c).value
                        if v and ("Transport cost" in str(v) or "cost" in str(v).lower()):
                            cost_name_rows.append(r)
                            break
                    else:
                        subheader_rows.append(r)
            break

    if header_row is None:
        return

    # The row after header is the value types row (Currency, Flat, p/unit)
    types_row = header_row + 1

    # Style cost name rows
    for r in cost_name_rows:
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=col)
            if cell.value:
                cell.font = COST_NAME_FONT

    # Style sub-header info rows (Applies if, Rate by, etc.)
    for r in subheader_rows:
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=col)
            cell.font = Font(italic=True, size=9, name="Calibri", color="595959")

    # Style the main header row
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # Style the types row
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=types_row, column=col)
        cell.font = Font(italic=True, size=9, name="Calibri")
        cell.fill = SUBHEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # Style data rows
    for row_idx in range(types_row + 1, ws.max_row + 1):
        first_val = ws.cell(row=row_idx, column=1).value
        if first_val is None:
            continue
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if col <= 6:
                cell.alignment = LEFT
            else:
                cell.alignment = CENTER

    # Freeze panes below header row
    ws.freeze_panes = ws.cell(row=types_row + 1, column=1)

    # Auto-width
    auto_width(ws, min_width=6, max_width=16)

    # Make first few columns (shipment info) wider
    for col_idx in range(1, min(7, ws.max_column + 1)):
        col_letter = get_column_letter(col_idx)
        current = ws.column_dimensions[col_letter].width
        ws.column_dimensions[col_letter].width = max(current, 14)


def format_accessorials_sheet(ws):
    """Format an accessorials sheet."""
    if ws.max_row < 2:
        return

    # Row 1 is the header
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = ACC_HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # Data rows
    for row_idx in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT if col <= 2 else CENTER

    # Freeze panes
    ws.freeze_panes = ws.cell(row=2, column=1)

    # Auto-width with wider columns for names
    auto_width(ws, min_width=8, max_width=40)

    # Rate Card cost Name column wider
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 25


def format_workbook(xlsx_path: Path):
    """Apply formatting to all sheets in a workbook."""
    wb = load_workbook(xlsx_path)

    for ws in wb.worksheets:
        sheet_name = ws.title.lower()
        if "acc" in sheet_name or "accessorial" in sheet_name:
            format_accessorials_sheet(ws)
        else:
            format_rates_sheet(ws)

    wb.save(xlsx_path)


def format_all_output(output_dir: Path):
    """Format all XLSX files in the output directory."""
    xlsx_files = list(output_dir.glob("*.xlsx"))
    for xlsx_path in xlsx_files:
        print(f"  Formatting: {xlsx_path.name}")
        format_workbook(xlsx_path)
    print(f"  Formatted {len(xlsx_files)} file(s)")
